import requests
import yaml
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import os
import base64
import json
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from openpyxl.styles import PatternFill
import argparse

appKey = "qpRMO6lj4smwkT1sWlSdIj7b8QF5rG8Q"
cruiselines = []
cruiselines.append({"lineName": "royalcaribbean", "lineCode": "R", "linePretty": "Royal Caribbean", "productList": ["beverage","dining","internet","onboardactivities","photoPackage","gifts","key","roomdelivery","packages","cococay","royalbeachclub","arcade","spa","preandpost"]})
cruiselines.append({"lineName": "celebritycruises", "lineCode": "C", "linePretty": "Celebrity", "productList": ["drinks","food","packages","shipexcursions","roomdelivery","spa","wifi","exclusiveexperiences","giftsandextras","preandpost","photoPackage","fitness","programming"]})

def main():
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(timestamp)
          
    parser = argparse.ArgumentParser()
    parser.add_argument("-b", "--booking", help="Booking ID") 
    args = parser.parse_args()

    bookingID = None
    if args.booking:
        bookingID = args.booking

    with open('config.yaml', 'r') as file:
        data = yaml.safe_load(file)
        
        currencyCode = 'USD'
        if 'currency' in data:
            currencyCode = data['currency']

        if 'accountInfo' in data:
            for accountInfo in data['accountInfo']:
                username = accountInfo['username']
                password = accountInfo['password']
                print(username, "Currency:", currencyCode) 
                session = requests.session()
                for cruiseline in cruiselines:
                    access_token,accountId,session = login(username,password,session,cruiseline['lineName'])
                    getVoyages(bookingID, access_token,accountId,session,cruiseline['lineCode'],cruiseline['productList'],currencyCode)
            
def login(username,password,session,cruiseLineName):
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Authorization': 'Basic ZzlTMDIzdDc0NDczWlVrOTA5Rk42OEYwYjRONjdQU09oOTJvMDR2TDBCUjY1MzdwSTJ5Mmg5NE02QmJVN0Q2SjpXNjY4NDZrUFF2MTc1MDk3NW9vZEg1TTh6QzZUYTdtMzBrSDJRNzhsMldtVTUwRkNncXBQMTN3NzczNzdrN0lC',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:136.0) Gecko/20100101 Firefox/136.0',
    }
    
    data = 'grant_type=password&username=' + username +  '&password=' + password + '&scope=openid+profile+email+vdsid'
    
    response = session.post('https://www.'+cruiseLineName+'.com/auth/oauth2/access_token', headers=headers, data=data)
    
    if response.status_code != 200:
        print(cruiseLineName + " Website Might Be Down. Quitting")
        quit()
          
    access_token = response.json().get("access_token")
    
    list_of_strings = access_token.split(".")
    string1 = list_of_strings[1]
    decoded_bytes = base64.b64decode(string1 + '==')
    auth_info = json.loads(decoded_bytes.decode('utf-8'))
    accountId = auth_info["sub"]
    return access_token,accountId,session

def getVoyages(bookingID,access_token,accountId,session,cruiseLineCode,productList,currencyCode):

    headers = {
        'Access-Token': access_token,
        'AppKey': appKey,
        'vds-id': accountId,
    }
        
    params = {
        'brand': cruiseLineCode,
        'includeCheckin': 'false',
    }

    response = requests.get(
        'https://aws-prd.api.rccl.com/v1/profileBookings/enriched/' + accountId,
        params=params,
        headers=headers,
    )

    for booking in response.json().get("payload").get("profileBookings"):
        if bookingID is not None and bookingID != booking.get("bookingId"):
            continue
        reservationId = booking.get("bookingId")
        passengerId = booking.get("passengerId")
        sailDate = booking.get("sailDate")
        numberOfNights = booking.get("numberOfNights")
        shipCode = booking.get("shipCode")
        getProducts(access_token,accountId,session,reservationId,passengerId,shipCode,sailDate,cruiseLineCode,productList,currencyCode)
        getOrders(access_token,accountId,session,reservationId,passengerId,shipCode,sailDate,numberOfNights,currencyCode)
        
def getProducts(access_token,accountId,session,reservationId,passengerId,ship,startDate,cruiseLineCode,productList,currencyCode):

    headers = {
        'Host': 'aws-prd.api.rccl.com',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:137.0) Gecko/20100101 Firefox/137.0',
        'Accept': 'application/json',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br, zstd',
        'X-Requested-With': 'XMLHttpRequest',
        'AppKey': appKey,
        'Content-Type': 'application/json',
        'Access-Token': access_token,
        'vds-id': accountId,
        'Account-Id': accountId,
        'X-Request-Id': '6830e53e32ff75f0ecac813a',
        'Req-App-Id': 'Celebrity.Web.PlanMyCruise',
        'Req-App-Vers': '1.76.2',
        'Origin': 'https://www.celebritycruises.com',
        'DNT': '1',
        'Sec-GPC': '1',
        'Connection': 'keep-alive',
        'Referer': 'https://www.celebritycruises.com/',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'cross-site',
        'Priority': 'u=0',
    }

    data = {
        "sortKey": "cRank-asc",
        "filterFacets": None
    }

    wbName = cruiseLineCode + "-" + reservationId + "-products-"+currencyCode+".xlsx"
    shName = str(datetime.now().strftime("%Y-%m-%d (%H %M)"))
    compPrice = [] # Array to compare prices to the previous value

    if os.path.isfile(wbName):
        workbook = openpyxl.load_workbook(wbName)
        for sheetname in workbook.sheetnames:
            if "Chart " in sheetname or sheetname == "Sheet":
                continue
            break
        lastSheet = workbook[sheetname]
        for i in range(1, lastSheet.max_row+1):
            compPrice.append({"key": lastSheet.cell(row=i, column=1).value + '|' + lastSheet.cell(row=i, column=2).value, "msrp": lastSheet.cell(row=i, column=3).value, "price": lastSheet.cell(row=i, column=4).value})
    else:
        workbook = openpyxl.Workbook()

    redFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
    greenFill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

    sheet = workbook.create_sheet(shName, 0)
    workbook.active = workbook[shName]
    currow = 2 # Header is row 1

    sheet.append(["Category","Title", "MSRP", "Price", "Unit", "Promo", "Pct", "Discount"])
    for cell in sheet["1:1"]:
        cell.style = "Headline 1"
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 53
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 33

    for product in productList:

        initialPageSize = 50

        postURL = 'https://aws-prd.api.rccl.com/en/celebrity/web/commerce-api/catalog/v2/'+ship+'/categories/'+product+'/products?reservationId='+reservationId+'&passengerId='+passengerId+'&startDate='+startDate+'&currentPage=0&pageSize='+str(initialPageSize)+'&currencyIso='+currencyCode+'&regionCode=EUROP'

        response = session.post(
            postURL,
            headers=headers,
            data=json.dumps(data)
        )
        
        if response.status_code != 200:
            print("getProducts initial - Status:" + str(response.status_code) + " Quitting")
            print(response.json())
            print(data)
            print(postURL)
            quit()

        # Make sure we have all the products (if there are more than default page size)
        num = response.json().get("payload").get("page").get("totalResults")
        if num > initialPageSize:
            response = session.post(
                'https://aws-prd.api.rccl.com/en/celebrity/web/commerce-api/catalog/v2/'+ship+'/categories/'+product+'/products?reservationId='+reservationId+'&passengerId='+passengerId+'&startDate='+startDate+'&currentPage=0&pageSize='+str(num)+'&currencyIso='+currencyCode+'&regionCode=EUROP',
                headers=headers,
                data=json.dumps(data)
            )
        
        if response.status_code != 200:
            print("getProducts again - Status:" + str(response.status_code) + " Quitting")
            print(response.json())
            quit()

        category = product
        if response.json().get("payload").get("merchandisingCategoryWithChildren") and response.json().get("payload").get("merchandisingCategoryWithChildren").get("categoryDisplayName"):
            category = response.json().get("payload").get("merchandisingCategoryWithChildren").get("categoryDisplayName")

        if response.json().get("payload").get("products"):
            for item in response.json().get("payload").get("products"):
                if item["stock"]["stockLevelStatus"] != "inStock" or item["title"] is None or item["lowestAdultPrice"] is None:
                    continue
                if item["promoDescription"] and item["promoDescription"]["displayName"]:
                    displayName = item["promoDescription"]["displayName"]
                else:
                    displayName = "None"
                if item["promoDescription"] and item["promoDescription"]["promotionValue"]:
                    promotionValue = int(item["promoDescription"]["promotionValue"]) / 100
                else:
                    promotionValue = ""
                if item["promoDescription"] and item["promoDescription"]["discountedValue"]:
                    discountedValue = item["promoDescription"]["discountedValue"]
                else:
                    discountedValue = 0
                if item["unit"] and item["unit"]["name"]:
                    displayUnit = item["unit"]["name"]
                else:
                    displayUnit = "None"
                if len(item["variantIdList"]) > 1:
                    # There are variants. Add row for each
                    for optn in item["variantIdList"]:
                        variant = getVariant(access_token,accountId,session,reservationId,passengerId,ship,startDate,product,optn,currencyCode)
                        sheet.append([category, item["title"] + " - " + variant["description"], variant["msrp"], variant["price"], displayUnit, variant["promoDescription"], variant["promotionValue"], variant["discountedValue"]])
                        sheet.cell(row=currow, column=3).number_format = '"$"#,##0.00'
                        sheet.cell(row=currow, column=4).number_format = '"$"#,##0.00'
                        sheet.cell(row=currow, column=7).number_format = FORMAT_PERCENTAGE
                        sheet.cell(row=currow, column=8).number_format = '"$"#,##0.00'
                        for element in compPrice:
                            if element["key"] == category + "|" + item["title"] and element["msrp"] == item["msrpAdultPrice"]:
                                if item["lowestAdultPrice"] > element["price"]:
                                    sheet.cell(row=currow, column=4).fill = redFill
                                    print(cruiseLineCode, reservationId, "Price increase:", category, item["title"], "from ${:0,.2f}".format(element["price"]), "to ${:0,.2f}".format(item["lowestAdultPrice"]))
                                elif item["lowestAdultPrice"] < element["price"]:
                                    sheet.cell(row=currow, column=4).fill = greenFill
                                    print(cruiseLineCode, reservationId, "Price decrease:", category, item["title"], "from ${:0,.2f}".format(element["price"]), "to ${:0,.2f}".format(item["lowestAdultPrice"]))
                                break
                        currow += 1
                else:
                    sheet.append([category, item["title"], item["msrpAdultPrice"], item["lowestAdultPrice"], displayUnit, displayName, promotionValue, discountedValue])
                    sheet.cell(row=currow, column=3).number_format = '"$"#,##0.00'
                    sheet.cell(row=currow, column=4).number_format = '"$"#,##0.00'
                    sheet.cell(row=currow, column=7).number_format = FORMAT_PERCENTAGE
                    sheet.cell(row=currow, column=8).number_format = '"$"#,##0.00'
                    for element in compPrice:
                        if element["key"] == category + "|" + item["title"] and element["msrp"] == item["msrpAdultPrice"]:
                            if item["lowestAdultPrice"] > element["price"]:
                                sheet.cell(row=currow, column=4).fill = redFill
                                print(cruiseLineCode, reservationId, "Price increase:", category, item["title"], "from ${:0,.2f}".format(element["price"]), "to ${:0,.2f}".format(item["lowestAdultPrice"]))
                            elif item["lowestAdultPrice"] < element["price"]:
                                sheet.cell(row=currow, column=4).fill = greenFill
                                print(cruiseLineCode, reservationId, "Price decrease:", category, item["title"], "from ${:0,.2f}".format(element["price"]), "to ${:0,.2f}".format(item["lowestAdultPrice"]))
                            break
                    currow += 1

    sheet.freeze_panes = 'A2'
    workbook.save(wbName)
    workbook.close()

def getOrders(access_token,accountId,session,reservationId,passengerId,ship,startDate,numberOfNights,currencyCode):
    
    headers = {
        'Access-Token': access_token,
        'AppKey': appKey,
        'Account-Id': accountId,
    }
    
    params = {
        'passengerId': passengerId,
        'reservationId': reservationId,
        'sailingId': ship + startDate,
        'currencyIso': currencyCode,
        'includeMedia': 'false',
    }
    
    response = requests.get(
        'https://aws-prd.api.rccl.com/en/royal/web/commerce-api/calendar/v1/' + ship + '/orderHistory',
        params=params,
        headers=headers,
    )
 
    # Check for my orders and orders others booked for me
    for order in response.json().get("payload").get("myOrders") + response.json().get("payload").get("ordersOthersHaveBookedForMe"):
        orderCode = order.get("orderCode")
        
        # Only get Valid Orders That Cost Money
        if order.get("orderTotals").get("total") > 0: 
            
            # Get Order Details
            response = requests.get(
                'https://aws-prd.api.rccl.com/en/royal/web/commerce-api/calendar/v1/' + ship + '/orderHistory/' + orderCode,
                params=params,
                headers=headers,
            )
                    
            for orderDetail in response.json().get("payload").get("orderHistoryDetailItems"):
                # check for cancelled status at item-level
                if orderDetail.get("guests")[0].get("orderStatus") == "CANCELLED":
                    continue
                order_title = orderDetail.get("productSummary").get("title")
                product = orderDetail.get("productSummary").get("baseId")
                prefix = orderDetail.get("productSummary").get("productTypeCategory").get("id")
                paidPrice = orderDetail.get("guests")[0].get("priceDetails").get("subtotal")
                # These packages report total price, must divide by number of days
                if orderDetail.get("productSummary").get("salesUnit") in [ 'PER_NIGHT', 'PER_DAY' ]:
                   paidPrice = round(paidPrice / numberOfNights,2)
                getCurrentPrice(access_token,accountId,session,reservationId,passengerId,ship,startDate,prefix,paidPrice,product,currencyCode)

def getCurrentPrice(access_token,accountId,session,reservationId,passengerId,ship,startDate,prefix,paidPrice,product,currencyCode):    
    
    headers = {
        'Access-Token': access_token,
        'AppKey': appKey,
        'vds-id': accountId,
    }

    params = {
        'reservationId': reservationId,
        'startDate': startDate,
        'passengerId': passengerId,
        'currencyIso': currencyCode,
    }

    response = session.get(
        'https://aws-prd.api.rccl.com/en/royal/web/commerce-api/catalog/v2/' + ship + '/categories/' + prefix + '/products/' + str(product),
        params=params,
        headers=headers,
    )
    
    title = response.json().get("payload").get("title")
    currentPrice = None
    try:
        currentPrice = response.json().get("payload").get("startingFromPrice").get("adultPromotionalPrice")
        if not currentPrice:
            currentPrice = response.json().get("payload").get("startingFromPrice").get("adultShipboardPrice")
    except:
        pass
    
    if currentPrice:
        text = reservationId + ": " + title + " - Paid price: {:0,.2f}".format(paidPrice) + " Current price: {:0,.2f}".format(currentPrice)
        if currentPrice < paidPrice:
            text += " - DOWN {:0,.2f}".format(paidPrice - currentPrice)
        elif currentPrice > paidPrice:
            text += " - UP {:0,.2f}".format(currentPrice - paidPrice)
        else:
            text += " - unchanged"
    else:
        text = reservationId + ": " + title + " - Paid price: {:0,.2f}".format(paidPrice) + " Current price not available. Product code: " + product

    print(text)

def getVariant(access_token,accountId,session,reservationId,passengerId,ship,startDate,prefix,product,currencyCode):    
    
    variant = { 'product': product, 'description': '', 'price': None, 'msrp': None, 'promotionValue': None, 'discountedValue': None, 'promoDescription': 'None' }

    headers = {
        'Access-Token': access_token,
        'AppKey': appKey,
        'vds-id': accountId,
    }

    params = {
        'reservationId': reservationId,
        'startDate': startDate,
        'passengerId': passengerId,
        'currencyIso': currencyCode,
    }

    response = session.get(
        'https://aws-prd.api.rccl.com/en/royal/web/commerce-api/catalog/v2/' + ship + '/categories/' + prefix + '/products/' + str(product),
        params=params,
        headers=headers,
    )

    variant["description"] = response.json().get("payload").get("baseOptions")[0].get("selected").get("variantOptionQualifiers")[0].get("value")
    try:
        variant["price"] = response.json().get("payload").get("startingFromPrice").get("adultPromotionalPrice")
        if not variant["price"]:
             variant["price"] = response.json().get("payload").get("startingFromPrice").get("adultShipboardPrice")
        variant["msrp"] = response.json().get("payload").get("startingFromPrice").get("adultShipboardPrice")
        variant["promoDescription"] = response.json().get("payload").get("promoDescription").get("displayName")
        variant["discountedValue"] = response.json().get("payload").get("promoDescription").get("discountedValue")
        variant["promotionValue"] = int(response.json().get("payload").get("promoDescription").get("promotionValue")) / 100
    except:
        pass

    return variant

if __name__ == "__main__":
    main()