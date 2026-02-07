import requests
import yaml
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import os
import base64
import json
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
import argparse
from babel.numbers import get_currency_precision, get_currency_symbol, format_currency

appKey = "qpRMO6lj4smwkT1sWlSdIj7b8QF5rG8Q"
cruiselines = []
cruiselines.append({"lineName": "royalcaribbean", "lineCode": "R", "linePretty": "Royal Caribbean"})
cruiselines.append({"lineName": "celebritycruises", "lineCode": "C", "linePretty": "Celebrity"})
currencyCode = 'USD'
localeCode = 'en_US'

def main():
    global cruiseLineName, cruiseLineCode
    global currencyCode, localeCode

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(timestamp)

    parser = argparse.ArgumentParser()
    parser.add_argument("-b", "--booking", help="Booking ID") 
    args = parser.parse_args()

    bookingID = None
    if args.booking:
        bookingID = args.booking
          
    with open('config.yaml', 'r') as file:
        data = yaml.safe_load(file)
        
        if 'currency' in data:
            currencyCode = data['currency']
        if 'locale' in data:
            localeCode = data['locale']

        if 'accountInfo' in data:
            for accountInfo in data['accountInfo']:
                username = accountInfo['username']
                password = accountInfo['password']
                print(username, "Currency:", currencyCode, "Locale:", localeCode) 
                session = requests.session()
                for cruiseline in cruiselines:
                    access_token,accountId,session = login(username,password,session,cruiseline['lineName'])
                    getVoyages(bookingID, access_token,accountId,session,cruiseline['lineCode'])
            
def login(username,password,session,cruiseLineName):
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Authorization': 'Basic ZzlTMDIzdDc0NDczWlVrOTA5Rk42OEYwYjRONjdQU09oOTJvMDR2TDBCUjY1MzdwSTJ5Mmg5NE02QmJVN0Q2SjpXNjY4NDZrUFF2MTc1MDk3NW9vZEg1TTh6QzZUYTdtMzBrSDJRNzhsMldtVTUwRkNncXBQMTN3NzczNzdrN0lC',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:136.0) Gecko/20100101 Firefox/136.0',
    }
    
    data = 'grant_type=password&username=' + username +  '&password=' + password + '&scope=openid+profile+email+vdsid'
    
    response = session.post('https://www.'+cruiseLineName+'.com/auth/oauth2/access_token', headers=headers, data=data)
    
    if response.status_code != 200:
        print(cruiseLineName + " Website Might Be Down. Quitting")
        quit()
          
    access_token = response.json().get("access_token")
    
    list_of_strings = access_token.split(".")
    string1 = list_of_strings[1]
    decoded_bytes = base64.b64decode(string1 + '==')
    auth_info = json.loads(decoded_bytes.decode('utf-8'))
    accountId = auth_info["sub"]
    return access_token,accountId,session

def getVoyages(bookingID,access_token,accountId,session,cruiseLineCode):

    headers = {
        'Access-Token': access_token,
        'AppKey': appKey,
        'vds-id': accountId,
    }
        
    params = {
        'brand': cruiseLineCode,
        'includeCheckin': 'false',
    }

    response = requests.get(
        'https://aws-prd.api.rccl.com/v1/profileBookings/enriched/' + accountId,
        params=params,
        headers=headers,
    )

    for booking in response.json().get("payload").get("profileBookings"):
        if bookingID is not None and bookingID != booking.get("bookingId"):
            continue
        reservationId = booking.get("bookingId")
        passengerId = booking.get("passengerId")
        sailDate = booking.get("sailDate")
        numberOfNights = booking.get("numberOfNights")
        shipCode = booking.get("shipCode")
        getProducts(access_token,accountId,session,reservationId,passengerId,shipCode,sailDate,cruiseLineCode)
        
def getProducts(access_token,accountId,session,reservationId,passengerId,ship,startDate,cruiseLineCode):

    headers = {
        'Host': 'aws-prd.api.rccl.com',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:137.0) Gecko/20100101 Firefox/137.0',
        'Accept': 'application/json',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br, zstd',
        'X-Requested-With': 'XMLHttpRequest',
        'AppKey': appKey,
        'Content-Type': 'application/json',
        'Access-Token': access_token,
        'vds-id': accountId,
        'Account-Id': accountId,
        'X-Request-Id': '6830e53e32ff75f0ecac813a',
        'Req-App-Id': 'Celebrity.Web.PlanMyCruise',
        'Req-App-Vers': '1.76.2',
        'Origin': 'https://www.celebritycruises.com',
        'DNT': '1',
        'Sec-GPC': '1',
        'Connection': 'keep-alive',
        'Referer': 'https://www.celebritycruises.com/',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'cross-site',
        'Priority': 'u=0',
    }

    data = {
        "sortKey": "cRank-asc",
        "filterFacets": None
    }

    response = session.post(
        'https://aws-prd.api.rccl.com/en/celebrity/web/commerce-api/catalog/v2/'+ship+'/categories/shorex/products?reservationId='+reservationId+'&passengerId='+passengerId+'&startDate='+startDate+'&currentPage=0&pageSize=1&currencyIso='+currencyCode+'&regionCode=EUROP',
        headers=headers,
        data=json.dumps(data)
    )
    
    if response.status_code != 200:
        print("getProducts - Status:" + str(response.status_code) + " Quitting")
        print(response.json())
        quit()

    # Make sure we have all the excursions
    num = response.json().get("payload").get("page").get("totalResults")
    response = session.post(
        'https://aws-prd.api.rccl.com/en/celebrity/web/commerce-api/catalog/v2/'+ship+'/categories/shorex/products?reservationId='+reservationId+'&passengerId='+passengerId+'&startDate='+startDate+'&currentPage=0&pageSize='+str(num)+'&currencyIso='+currencyCode+'&regionCode=EUROP',
        headers=headers,
        data=json.dumps(data)
    )
    
    if response.status_code != 200:
        print("getProducts - Status:" + str(response.status_code) + " Quitting")
        print(response.json())
        quit()

    wbName = cruiseLineCode+"-"+reservationId+"-shorex-"+currencyCode+".xlsx"
    shName = str(datetime.now().strftime("%Y-%m-%d (%H %M)"))

    if os.path.isfile(wbName):
        workbook = openpyxl.load_workbook(wbName)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.create_sheet(shName, 0)
    workbook.active = workbook[shName]
    currow = 2 # Header is row 1

    sheet.append(["Link", "Day", "Title", "Port", "Time", "Duration", "Price", "Promotion", "Description" ])
    for cell in sheet["1:1"]:
        cell.style = "Headline 1"
    sheet.column_dimensions['A'].width = 7
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 53
    sheet.column_dimensions['D'].width = 27
    sheet.column_dimensions['E'].width = 9
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 11
    sheet.column_dimensions['H'].width = 24
    sheet.column_dimensions['I'].width = 60

    for excursion in response.json().get("payload").get("products"):
        getURL = 'https://aws-prd.api.rccl.com//en/celebrity/web/commerce-api/catalog/v2/'+ship+'/categories/pt_shoreX/products/'+excursion["id"]+'?reservationId='+reservationId+'&passengerId='+passengerId+'&startDate='+startDate+'&currencyIso='+currencyCode

        detail = session.get(getURL, headers=headers)
        
        if detail.status_code != 200:
            print("getProducts - Status:" + str(detail.status_code) + " Quitting")
            print(detail.json())
            print(getURL)
            quit()

        for offering in detail.json().get("payload").get("bookingOfferingData").get("offerings"):
            if detail.json().get("payload").get("bookingEligibility").get("allowed") and excursion["lowestAdultPrice"] > 0:
                duration = ""
                if detail.json().get("payload").get("durationValues") is not None:
                    duration = detail.json().get("payload").get("durationValues")[0]
                if excursion["promoDescription"] and excursion["promoDescription"]["displayName"]:
                    displayName = excursion["promoDescription"]["displayName"]
                elif excursion["promoDescription"] and excursion["promoDescription"]["title"]:
                    displayName = excursion["promoDescription"]["title"]
                else:
                    displayName = "None"
                sheet.append(["", offering["dayOfCruise"], excursion["title"], offering["portLocation"], offering["dateTime"], duration, excursion["lowestAdultPrice"], displayName, detail.json().get("payload").get("detail").replace('<p>', '').replace('</p>','')])
                link = "https://www.celebritycruises.com/account/cruise-planner/category/pt_shoreX/product/"+excursion["id"]+"?bookingId="+reservationId+"&shipCode="+ship+"&sailDate="+startDate
                sheet.cell(row=currow, column=1).value = excursion["id"]
                sheet.cell(row=currow, column=1).hyperlink = link
                sheet.cell(row=currow, column=1).style = "Hyperlink"
                if offering["dateTime"] is not None:
                    offer_date = datetime.strptime(offering["dateTime"], '%Y-%m-%dT%H:%M:%S')
                    sheet.cell(row=currow, column=5).value = offer_date
                    sheet.cell(row=currow, column=5).number_format = 'h:mm AM/PM'
                else:
                    sheet.cell(row=currow, column=5).value = ""
                sheet.cell(row=currow, column=7).number_format = excel_currency_format(currencyCode,localeCode)
                for col in range(1, sheet.max_column+1):
                    sheet.cell(row=currow, column=col).alignment = Alignment(vertical='top')
                sheet.cell(row=currow, column=9).alignment = Alignment(wrap_text=True,vertical='top')
                currow += 1

    sheet.freeze_panes = 'A2'
    workbook.save(wbName)

def excel_currency_format(currency, locale):
    symbol = get_currency_symbol(currency, locale=locale)
    decimals = get_currency_precision(currency)
    decimal_part = "." + ("0" * decimals) if decimals else ""
    return f'"{symbol}"#,##0{decimal_part}'

def format_money(amount, currency, locale):
    return format_currency(amount, currency, locale=locale)

if __name__ == "__main__":
    main()