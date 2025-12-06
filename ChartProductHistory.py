import openpyxl
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils import get_column_letter
import argparse
from datetime import datetime

def build_chart_from_description(filename, description):
    wb = openpyxl.load_workbook(filename)

    results = []
    min_amt = None
    max_amt = None
    prev_amt = 0
    prev_cnt = 0
    
    for sheetname in wb.sheetnames:
        if "Chart " in sheetname or sheetname == "Sheet":
            continue
        sheet = wb[sheetname]
        date = datetime.strptime(sheetname, "%Y-%m-%d (%H %M)")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            desc = row[1]
            amount = row[3]
            if desc and description.lower() in str(desc).lower():
                if prev_amt is None:
                    prev_amt = amount
                if amount != prev_amt:
                    prev_amt = amount
                    prev_cnt += 1
                if max_amt is None or amount > max_amt:
                    max_amt = amount
                if min_amt is None or amount < min_amt:
                    min_amt = amount
                results.append((date, amount))

    if results:
        print("Minimum price", "${:0,.2f}".format(min_amt), "Maximum price", "${:0,.2f}".format(max_amt),"Price changes", str(prev_cnt))
        results.sort(key=lambda x: x[0])

        chart_sheet_name = "Chart " + description[:24]
        if chart_sheet_name in wb.sheetnames:
            chart_sheet = wb[chart_sheet_name]
            wb.remove(chart_sheet)
        chart_sheet = wb.create_sheet(chart_sheet_name, 0)

        chart_sheet.append(["Date", "Amount"])
        for date, amount in results:
            chart_sheet.append([date, amount])

        chart = BarChart()
        chart.title = description
        chart.x_axis.title = "Date"
        chart.legend.position = "b"
        chart.y_axis.title = "Amount"
        chart.y_axis.scaling.min = min_amt - 20 # minimum limit of axis
        chart.y_axis.scaling.max = max_amt + 20 # maximum limit if axis
        chart.y_axis.majorUnit = 20   # spacing of gridlines

        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(results)+1)
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(results)+1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart_sheet.add_chart(chart, "E5")

        wb.save(filename)

parser = argparse.ArgumentParser()
parser.add_argument("-f", "--filename", help="filename") 
parser.add_argument("-p", "--product", help="Product") 
args = parser.parse_args()

build_chart_from_description(args.filename, args.product)
