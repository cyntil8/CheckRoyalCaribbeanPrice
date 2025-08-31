import requests
import yaml
from apprise import Apprise
from datetime import datetime
from bs4 import BeautifulSoup
from urllib.parse import urlparse, parse_qs
import re
import base64
import json
import os
import sys
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Series, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

appKey = "hyNNqIPHHzaLzVpcICPdAdbFV8yvTsAm"
cruiselines = []
cruiselines.append({"lineName": "royalcaribbean", "lineCode": "R", "linePretty": "Royal Caribbean"})
cruiselines.append({"lineName": "celebritycruises", "lineCode": "C", "linePretty": "Celebrity"})

def main():
    timestamp = datetime.now()
    print(timestamp.strftime("%Y-%m-%d %H:%M:%S"))
            
    with open('config.yaml', 'r') as file:
        data = yaml.safe_load(file)
        
        if 'cruises' in data:
            for cruiseline in cruiselines:
                print("Checking prices for your " + cruiseline['linePretty'] + " cruises")
                for cruises in data['cruises']:
                    if cruiseline['lineName'] in cruises['cruiseURL']:
                        compPrice = float(cruises['paidPrice'])
                        get_cruise_price(timestamp, cruises['cruiseURL'], compPrice, cruiseline['lineName'])
            
def get_cruise_price(timestamp, url, compPrice, cruiseLineName):

    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-US,en;q=0.9',
        'priority': 'u=0, i',
        'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Google Chrome";v="134"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36',
    }

    parsed_url = urlparse(url)
    params = parse_qs(parsed_url.query)
    params.pop('r0y', None)
    params.pop('r0x', None)

    response = requests.get('https://www.'+cruiseLineName+'.com/checkout/guest-info', params=params, headers=headers)
    
    preString = params.get("sailDate")[0] + " " + params.get("shipCode")[0]+ " " + params.get("r0d")[0] + " " + params.get("r0f")[0]
    
    soup = BeautifulSoup(response.text, "html.parser")
    soupFind = soup.find("span",attrs={"class":"SummaryPrice_title__1nizh9x5","data-testid":"pricing-total"})
    if soupFind is None:
        m = re.search("\"B:0\",\"NEXT_REDIRECT;replace;(.*);307;", response.text)
        if m is not None:
            redirectString = m.group(1)
            textString = preString + ": URL Not Working - Redirecting to suggested room"
            print(textString)
            newURL = "https://www." + cruiseLineName + ".com" + redirectString
            get_cruise_price(timestamp, newURL, compPrice, cruiseLineName)
            print("Update url to: " + newURL)
            return
        else:
            textString = preString + " No Longer Available To Book"
            print(textString)
            return
    
    priceString = soupFind.text
    priceString = priceString.replace(",", "")
    m = re.search("\\$(.*)USD", priceString)
    priceOnlyString = m.group(1)
    currentPrice = float(priceOnlyString)
    
    soupFind = soup.find("span",attrs={"class":"ItineraryDetails_itemContent__g7io8o4 typography_fontLabelSmallReg__44ku1hn","data-testid":"itinerary-summary-ship"})
    if soupFind:
        shipName = soupFind.text
    else:
        shipName = params.get("shipCode")[0]
    soupFind = soup.find("button",attrs={"data-testid":"navigation-card-room-type-link"})
    if soupFind:
        roomType = soupFind.text
    else:
        roomType = params.get("r0f")[0]
    soupFind = soup.find("span",attrs={"data-testid":"itinerary-summary-start-date"})
    if soupFind:
        shipDate = soupFind.text
    else:
        shipDate = params.get("sailDate")[0][5:7] + "/" + params.get("sailDate")[0][8:10] + "/" + params.get("sailDate")[0][:4]
    group = shipDate + " " + shipName + ", " + roomType

    textString = preString + ": Saved Price {:0,.2f}".format(compPrice) + " Current Price {:0,.2f}".format(currentPrice)
    if currentPrice < compPrice: 
        textString += " - DOWN {:0,.2f}".format(compPrice - currentPrice)
    elif currentPrice > compPrice:
        textString += " - UP {:0,.2f}".format(currentPrice - compPrice)
    else:
        textString += " - unchanged"

    print(textString)

    wbName = "price_history.xlsx"
  
    if os.path.isfile(wbName):
        workbook = openpyxl.load_workbook(wbName)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Checked", group])
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 40

    addNewRow = addNewColumn = True

    # Look for a row with the timestamp. If not found, add it
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=1).value == timestamp.strftime("%Y-%m-%d %H:%M:%S"):
            addNewRow = False
            break

    if addNewRow:
        i = sheet.max_row + 1
        sheet.cell(row=i, column=1).value = timestamp.strftime("%Y-%m-%d %H:%M:%S")

    # Look for a column with the cruise. If not found, add it
    for x in range(2, sheet.max_column+1):
        if sheet.cell(row=1, column=x).value == group:
            addNewColumn = False
            break

    if addNewColumn:
        x = sheet.max_column + 1
        sheet.cell(row=1, column=x).value = group
        sheet.cell(row=2, column=x).value = currentPrice # Set base price
        sheet.column_dimensions[get_column_letter(x)].width = 40

    sheet.cell(row=i, column=x).value = currentPrice
    sheet.cell(row=i, column=x).number_format = '"$"##,##0.00'

    redFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
    greenFill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
    noFill = PatternFill(fill_type=None)

    # Mark the lowest price
    lowPriceRow = highPriceRow = 2
    lowestPrice = highestPrice = sheet.cell(row=2, column=x).value
    for row in range(3, sheet.max_row+1):
        if sheet.cell(row=row, column=x).value is not None:
            sheet.cell(row=row, column=x).fill = noFill
            if sheet.cell(row=row, column=x).value < lowestPrice:
                lowestPrice = sheet.cell(row=row, column=x).value
                lowPriceRow = row
            if sheet.cell(row=row, column=x).value > highestPrice:
                highestPrice = sheet.cell(row=row, column=x).value
                highPriceRow = row

    for row in range(2, sheet.max_row+1):
        if sheet.cell(row=row, column=x).value is not None:
            if sheet.cell(row=row, column=x).value == lowestPrice:
                sheet.cell(row=row, column=x).fill = greenFill
            if highestPrice > lowestPrice:
                if sheet.cell(row=row, column=x).value == highestPrice:
                    sheet.cell(row=row, column=x).fill = redFill

    sheet.freeze_panes = 'A2'

    # Delete any current chart
    if sheet._charts and sheet._charts[0]:
        del sheet._charts[0]

    # Add new chart
    chart = LineChart()
    chart.title = "Cruise Price Trends"
    chart.style = 2
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.line.width = 50000  # Very thick outline    
    chart.y_axis.title = "Price"
    chart.y_axis.numFmt = '"$"##,##0.00'
    chart.x_axis.title = "Checked"

    data = Reference(sheet, min_col=2, min_row=1, max_col=sheet.max_column, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend.position = "b"

    sheet.add_chart(chart, "A" + str(sheet.max_row + 3))

    workbook.save(wbName)
    workbook.close()

if __name__ == "__main__":
    main()